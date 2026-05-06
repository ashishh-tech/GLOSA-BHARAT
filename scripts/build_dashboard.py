import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

print("=" * 60)
print("  GLOSA-BHARAT | DATA DRIFT DASHBOARD GENERATOR")
print("=" * 60)

# ============================================================
# PHASE 1: LOAD & CLEAN
# ============================================================
print("\n[1/6] Loading and cleaning data...")

df = pd.read_csv(r'c:\Users\name\Desktop\PROJECTS\GLOSA-BHARAT-main\highway dataset.csv')
raw_count = len(df)
print(f"  Raw rows: {raw_count}")

df = df.drop_duplicates()
print(f"  After deduplication: {len(df)} rows  (removed: {raw_count - len(df)})")

df['vehicles'] = pd.to_numeric(df['vehicles'], errors='coerce')
df['day of week'] = df['day of week'].str.strip().str.title()
df['congestion']  = df['congestion'].str.strip().str.title()

df['vehicles']        = df.groupby('road segment id')['vehicles'].transform(lambda x: x.fillna(x.median()))
df['traffic density'] = df.groupby('road segment id')['traffic density'].transform(lambda x: x.fillna(x.median()))
df['avg speed kmph']  = df.groupby('road segment id')['avg speed kmph'].transform(lambda x: x.fillna(x.mean()))
df['CO2 emission']    = df.groupby('congestion')['CO2 emission'].transform(lambda x: x.fillna(x.mean()))
df['NOx']    = df['NOx'].fillna(df['NOx'].mean())
df['PM2.5']  = df['PM2.5'].fillna(df['PM2.5'].mean())
df['rain mm']= df['rain mm'].fillna(0)
df['visibility m'] = df['visibility m'].fillna(df['visibility m'].mean())

print(f"  Blanks imputed. Final clean rows: {len(df)}")

# ============================================================
# PHASE 2: FEATURE ENGINEERING
# ============================================================
print("\n[2/6] Engineering features...")
df['timestamp'] = pd.to_datetime(df['timestamp'])
df['hour'] = df['timestamp'].dt.hour
df['traffic efficiency index'] = (
    df['avg speed kmph'] / df['traffic density'].replace(0, np.nan)
).replace([np.inf, -np.inf], np.nan).fillna(df['avg speed kmph'] / df['traffic density'].median())
print("  Added: hour, traffic efficiency index")

# ============================================================
# PHASE 3: ANALYSES
# ============================================================
print("\n[3/6] Running all analyses...")

peak_hours  = df.groupby('hour')['vehicles'].sum().reset_index()
peak_hours.columns = ['Hour', 'Total Vehicles']

speed_seg = df.groupby('road segment id')['avg speed kmph'].mean().reset_index()
speed_seg.columns = ['Segment', 'Avg Speed (kmph)']
speed_seg = speed_seg.sort_values('Avg Speed (kmph)').reset_index(drop=True)

density_seg = df.groupby('road segment id')['traffic density'].mean().reset_index()
density_seg.columns = ['Segment', 'Avg Density']
density_seg = density_seg.sort_values('Avg Density', ascending=False).reset_index(drop=True)

co2_cong = df.groupby('congestion')['CO2 emission'].mean().reset_index()
co2_cong.columns = ['Congestion Level', 'Avg CO2']
omap = {'Low': 0, 'Medium': 1, 'High': 2}
co2_cong['ord'] = co2_cong['Congestion Level'].map(omap)
co2_cong = co2_cong.sort_values('ord').drop('ord', axis=1).reset_index(drop=True)

bins   = [-1, 0, 10, 25, 100]
labels = ['Dry (0mm)', 'Light (1-10)', 'Moderate (11-25)', 'Heavy (>25)']
df['rain_bin'] = pd.cut(df['rain mm'], bins=bins, labels=labels)
rain_speed = df.groupby('rain_bin', observed=True)['avg speed kmph'].mean().reset_index()
rain_speed.columns = ['Rainfall', 'Avg Speed (kmph)']

top5 = df.nlargest(5, 'CO2 emission')[
    ['x-coord', 'y-coord', 'CO2 emission', 'road segment id', 'congestion']
].reset_index(drop=True)
top5.index += 1

gore = df.groupby('gore point').agg(
    Avg_Speed=('avg speed kmph', 'mean'),
    Avg_CO2=('CO2 emission', 'mean'),
    Avg_Density=('traffic density', 'mean')
).reset_index()
gore['gore point'] = gore['gore point'].map({0: 'Non-Gore', 1: 'Gore Point'})

base_co2 = df['CO2 emission'].mean()
curr_ev   = df['ev percentage'].mean()
ev_sim = pd.DataFrame({
    'EV Scenario': ['Current', '25%', '50%', '75%', '100%'],
    'Avg CO2':     [base_co2*(1 - e*0.85) for e in [curr_ev, 0.25, 0.50, 0.75, 1.00]],
    'CO2 Reduc%':  [round((1-(1-e*0.85))*100, 1) for e in [curr_ev, 0.25, 0.50, 0.75, 1.00]]
})

tei_rank = df.groupby('road segment id')['traffic efficiency index'].mean().reset_index()
tei_rank.columns = ['Segment', 'Avg TEI']
tei_rank = tei_rank.sort_values('Avg TEI', ascending=False).reset_index(drop=True)
tei_rank['Rank'] = range(1, len(tei_rank)+1)

day_order = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
day_veh = df.groupby('day of week')['vehicles'].sum()
day_veh = day_veh.reindex([d for d in day_order if d in day_veh.index]).reset_index()
day_veh.columns = ['Day', 'Total Vehicles']

cong_dist = df['congestion'].value_counts().reset_index()
cong_dist.columns = ['Congestion', 'Count']

# KPIs
total_vol  = int(df['vehicles'].sum())
avg_spd    = df['avg speed kmph'].mean()
avg_co2    = df['CO2 emission'].mean()
em_per_veh = (df['CO2 emission'] / df['vehicles']).mean()
avg_tei    = df['traffic efficiency index'].mean()
total_nox  = df['NOx'].sum()
total_pm25 = df['PM2.5'].sum()

peak_hr     = int(peak_hours.loc[peak_hours['Total Vehicles'].idxmax(), 'Hour'])
slowest_seg = speed_seg.iloc[0]['Segment']
fastest_seg = speed_seg.iloc[-1]['Segment']
dense_seg   = density_seg.iloc[0]['Segment']
top_tei_seg = tei_rank.iloc[0]['Segment']
co2_drop50  = ev_sim[ev_sim['EV Scenario']=='50%']['CO2 Reduc%'].values[0]

print("  All analyses complete!")

# ============================================================
# PHASE 4: EXCEL WORKBOOK
# ============================================================
print("\n[4/6] Building Excel workbook...")

wb  = Workbook()
wb.remove(wb.active)

# ---- STYLE HELPERS ----
def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def ft(bold=False, size=11, color="FFFFFF", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def al(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(col="333333"):
    s = Side(style='thin', color=col)
    return Border(left=s, right=s, top=s, bottom=s)

# Colour palette
BG      = "0D1117"
CARD    = "161B22"
BLUE    = "3D6EFF"
TEAL    = "00C9A7"
RED     = "FF6B6B"
YELLOW  = "FFD166"
PURPLE  = "C084FC"
HEADER  = "0D3B5E"
GREY    = "8B949E"
DARK2   = "0A1520"

def cell_write(ws, row, col, value, bg=None, font=None, align=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:      c.fill      = fill(bg)
    if font:    c.font      = font
    if align:   c.alignment = align
    if num_fmt: c.number_format = num_fmt
    return c

def merge_cell(ws, r1, c1, r2, c2, value="", bg=None, font=None, align=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    if bg:    c.fill      = fill(bg)
    if font:  c.font      = font
    if align: c.alignment = align
    return c

# ============================================================
# SHEET A: ANALYSIS DATA  (source for charts)
# ============================================================
ws_ad = wb.create_sheet("Analysis Data")

def write_table(ws, start_row, start_col, dataframe, title, hdr_color=BLUE):
    # Title row
    ncols = len(dataframe.columns)
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=start_row, end_column=start_col+ncols-1)
    tc = ws.cell(row=start_row, column=start_col, value=title)
    tc.fill = fill(hdr_color); tc.font = ft(bold=True, size=11)
    tc.alignment = al('center')
    # Header
    for ci, col in enumerate(dataframe.columns):
        hc = ws.cell(row=start_row+1, column=start_col+ci, value=col)
        hc.fill = fill("1C2A3A")
        hc.font = ft(bold=True, size=10, color="AADDFF")
        hc.alignment = al('center')
        ws.column_dimensions[get_column_letter(start_col+ci)].width = 22
    # Rows
    for ri, row_vals in enumerate(dataframe.values):
        bg = DARK2 if ri % 2 == 0 else BG
        for ci, val in enumerate(row_vals):
            dc = ws.cell(row=start_row+2+ri, column=start_col+ci,
                         value=round(float(val), 3) if isinstance(val, (float, np.floating)) else val)
            dc.fill = fill(bg)
            dc.font = ft(size=10, color="DDDDDD")
            dc.alignment = al('center')
    return start_row + 2 + len(dataframe) + 1   # next free row

# Write all tables into Analysis Data sheet
r = 1
tables = [
    (peak_hours,   "A1: Peak Traffic Hours",              BLUE),
    (speed_seg,    "A2: Avg Speed by Segment",            "6B4EFF"),
    (density_seg,  "A3: Avg Density by Segment",          RED),
    (co2_cong,     "A4: CO2 by Congestion Level",         "E6851F"),
    (rain_speed,   "A5: Rainfall vs Avg Speed",           TEAL),
    (top5,         "A6: Top 5 Emission Hotspots",         RED),
    (gore,         "ADV1: Gore Point Impact",             YELLOW),
    (ev_sim,       "ADV2: EV Simulation",                 TEAL),
    (tei_rank,     "ADV3: TEI Ranking",                   BLUE),
    (day_veh,      "BONUS: Volume by Day of Week",        PURPLE),
    (cong_dist,    "BONUS2: Congestion Distribution",     "E6851F"),
]

# Write all in col 1
for df_t, title, color in tables:
    r = write_table(ws_ad, r, 1, df_t, title, color)

# ============================================================
# Helper: locate a table by title in ws_ad
# ============================================================
def find_table(ws, title):
    for row in ws.iter_rows(max_col=10, max_row=500):
        for c in row:
            if c.value == title:
                r0 = c.row    # title row
                r_hdr  = r0 + 1
                r_data = r0 + 2
                c_start = c.column
                # find data end row
                r_end = r_data
                while ws.cell(row=r_end, column=c_start).value is not None:
                    r_end += 1
                r_end -= 1
                # find col end
                c_end = c_start
                while ws.cell(row=r_hdr, column=c_end).value is not None:
                    c_end += 1
                c_end -= 1
                return r_hdr, r_data, r_end, c_start, c_end
    return None

# ============================================================
# SHEET B: DASHBOARD
# ============================================================
ws_d = wb.create_sheet("Dashboard", 0)

# Make every cell dark background
for r in range(1, 150):
    ws_d.row_dimensions[r].height = 15
    for c in range(1, 36):
        ws_d.cell(row=r, column=c).fill = fill(BG)

for c in range(1, 36):
    ws_d.column_dimensions[get_column_letter(c)].width = 6.2

# ---- TITLE BANNER ----
merge_cell(ws_d, 1, 1, 3, 35,
    "  SMART HIGHWAY TRAFFIC ANALYSIS DASHBOARD  |  SRIJAN '26 — DATA DRIFT",
    bg=HEADER,
    font=ft(bold=True, size=16, color=TEAL),
    align=al('center'))
ws_d.row_dimensions[1].height = 28
ws_d.row_dimensions[2].height = 28
ws_d.row_dimensions[3].height = 28

merge_cell(ws_d, 4, 1, 4, 35,
    f"   Clean Records: {len(df):,}  |  Road Segments: 8  |  Period: Jan–Mar 2026  |  Duplicates Removed: {raw_count - len(df)}",
    bg=BG,
    font=ft(size=10, color=GREY, italic=True),
    align=al('left', 'center'))

# Divider
for c in range(1, 36):
    ws_d.cell(row=5, column=c).fill = fill(BLUE)
ws_d.row_dimensions[5].height = 3

# ---- KPI CARDS (row 6–13) ----
kpi_cards = [
    ("TOTAL VEHICLES",      f"{total_vol:,.0f}",    "sum",       BLUE,   1,  7),
    ("AVG SPEED",           f"{avg_spd:.1f} kmph",  "all segs",  TEAL,   8, 14),
    ("AVG CO2 EMISSION",    f"{avg_co2:.1f}",        "units",     RED,   15, 21),
    ("EMISSION/VEHICLE",    f"{em_per_veh:.5f}",    "ratio",     YELLOW, 22, 28),
    ("AVG EFFICIENCY (TEI)",f"{avg_tei:.4f}",       "speed/density", PURPLE, 29, 35),
]

for title, value, unit, accent, c1, c2 in kpi_cards:
    for rr in range(6, 14):
        for cc in range(c1, c2+1):
            ws_d.cell(row=rr, column=cc).fill = fill(CARD)
    # accent bar top
    for cc in range(c1, c2+1):
        ws_d.cell(row=6, column=cc).fill = fill(accent)
    ws_d.row_dimensions[6].height = 6
    merge_cell(ws_d, 7, c1, 8, c2, title, bg=CARD,
               font=ft(bold=True, size=8, color=GREY), align=al('center'))
    merge_cell(ws_d, 9, c1, 11, c2, value, bg=CARD,
               font=ft(bold=True, size=16, color=accent), align=al('center'))
    merge_cell(ws_d, 12, c1, 13, c2, unit, bg=CARD,
               font=ft(size=8, italic=True, color=GREY), align=al('center'))

# Divider
for c in range(1, 36):
    ws_d.cell(row=14, column=c).fill = fill(BLUE)
ws_d.row_dimensions[14].height = 3

# ---- CHART SECTION LABELS ----
def section_header(ws, row, c1, c2, text, color=TEAL):
    merge_cell(ws, row, c1, row, c2, f"  {text}",
               bg=BG,
               font=ft(bold=True, size=11, color=color),
               align=al('left', 'center'))
    ws.row_dimensions[row].height = 20

# ============================================================
# SECTION HEADERS on Dashboard
# ============================================================
section_header(ws_d, 15, 1, 17, "PEAK TRAFFIC HOURS", BLUE)
section_header(ws_d, 15, 19, 35, "AVG SPEED BY SEGMENT", "6B4EFF")
section_header(ws_d, 50, 1, 17, "CO2 EMISSIONS BY CONGESTION", RED)
section_header(ws_d, 50, 19, 35, "RAINFALL VS SPEED", TEAL)
section_header(ws_d, 85, 1, 17, "DAILY TRAFFIC VOLUME (BONUS)", YELLOW)
section_header(ws_d, 85, 19, 35, "SEGMENT TEI RANKING (ADV)", PURPLE)
section_header(ws_d, 120, 1, 17, "EV SIMULATION -> CO2 DROP (ADV)", TEAL)
section_header(ws_d, 120, 19, 35, "GORE POINT SPEED IMPACT (ADV)", YELLOW)

# ============================================================
# BUILD & PLACE CHARTS
# ============================================================
print("\n[5/6] Building charts...")

def bar_chart(ws_data, cat_r, cat_c, val_r, val_c, nrows, title, y_label, color="4472C4",
              w=14, h=11, has_title_row=True):
    ch = BarChart()
    ch.type = "col"
    ch.title = title
    ch.y_axis.title = y_label
    ch.x_axis.title = None
    ch.style  = 10
    ch.width  = w
    ch.height = h
    cats = Reference(ws_data, min_col=cat_c, max_col=cat_c,
                     min_row=cat_r, max_row=cat_r+nrows-1)
    vals = Reference(ws_data, min_col=val_c, max_col=val_c,
                     min_row=val_r-1 if has_title_row else val_r,
                     max_row=val_r+nrows-1)
    ch.add_data(vals, titles_from_data=has_title_row)
    ch.set_categories(cats)
    try:
        ch.series[0].graphicalProperties.solidFill = color
        ch.series[0].graphicalProperties.line.solidFill = color
    except Exception:
        pass
    return ch

def line_chart(ws_data, cat_r, cat_c, val_r, val_c, nrows, title, y_label,
               color="00C9A7", w=14, h=11):
    ch = LineChart()
    ch.title = title
    ch.y_axis.title = y_label
    ch.style  = 10
    ch.width  = w
    ch.height = h
    cats = Reference(ws_data, min_col=cat_c, max_col=cat_c,
                     min_row=cat_r, max_row=cat_r+nrows-1)
    vals = Reference(ws_data, min_col=val_c, max_col=val_c,
                     min_row=val_r-1, max_row=val_r+nrows-1)
    ch.add_data(vals, titles_from_data=True)
    ch.set_categories(cats)
    try:
        ch.series[0].graphicalProperties.line.solidFill = color
        ch.series[0].graphicalProperties.line.width = 25000
        ch.series[0].marker.symbol = "circle"
        ch.series[0].marker.size   = 7
    except Exception:
        pass
    return ch

# Chart 1: Peak Traffic Hours
t = find_table(ws_ad, "A1: Peak Traffic Hours")
if t:
    rh, rd, re, c0, ce = t
    nrows = re - rd + 1
    ch1 = bar_chart(ws_ad, rd, c0, rd, c0+1, nrows,
                    "Peak Traffic Hours", "Total Vehicles", BLUE, w=14, h=11)
    ws_d.add_chart(ch1, "A16")

# Chart 2: Speed by Segment
t = find_table(ws_ad, "A2: Avg Speed by Segment")
if t:
    rh, rd, re, c0, ce = t
    nrows = re - rd + 1
    ch2 = bar_chart(ws_ad, rd, c0, rd, c0+1, nrows,
                    "Avg Speed per Segment", "Speed (kmph)", "6B4EFF", w=14, h=11)
    ws_d.add_chart(ch2, "S16")

# Chart 3: CO2 by Congestion
t = find_table(ws_ad, "A4: CO2 by Congestion Level")
if t:
    rh, rd, re, c0, ce = t
    nrows = re - rd + 1
    ch3 = bar_chart(ws_ad, rd, c0, rd, c0+1, nrows,
                    "Avg CO2 by Congestion Level", "Avg CO2", RED, w=14, h=11)
    ws_d.add_chart(ch3, "A51")

# Chart 4: Rainfall vs Speed
t = find_table(ws_ad, "A5: Rainfall vs Avg Speed")
if t:
    rh, rd, re, c0, ce = t
    nrows = re - rd + 1
    ch4 = line_chart(ws_ad, rd, c0, rd, c0+1, nrows,
                     "Rainfall Impact on Speed", "Avg Speed (kmph)", TEAL, w=14, h=11)
    ws_d.add_chart(ch4, "S51")

# Chart 5: Day of Week
t = find_table(ws_ad, "BONUS: Volume by Day of Week")
if t:
    rh, rd, re, c0, ce = t
    nrows = re - rd + 1
    ch5 = bar_chart(ws_ad, rd, c0, rd, c0+1, nrows,
                    "Traffic Volume by Day", "Total Vehicles", YELLOW, w=14, h=11)
    ws_d.add_chart(ch5, "A86")

# Chart 6: TEI Ranking
t = find_table(ws_ad, "ADV3: TEI Ranking")
if t:
    rh, rd, re, c0, ce = t
    nrows = re - rd + 1
    ch6 = bar_chart(ws_ad, rd, c0, rd, c0+1, nrows,
                    "Segment Efficiency (TEI)", "Avg TEI", PURPLE, w=14, h=11)
    ws_d.add_chart(ch6, "S86")

# Chart 7: EV Simulation
t = find_table(ws_ad, "ADV2: EV Simulation")
if t:
    rh, rd, re, c0, ce = t
    nrows = re - rd + 1
    ch7 = line_chart(ws_ad, rd, c0, rd, c0+1, nrows,
                     "EV Adoption -> CO2 Reduction", "Avg CO2", TEAL, w=14, h=11)
    ws_d.add_chart(ch7, "A121")

# Chart 8: Gore Point Speed
t = find_table(ws_ad, "ADV1: Gore Point Impact")
if t:
    rh, rd, re, c0, ce = t
    nrows = re - rd + 1
    ch8 = bar_chart(ws_ad, rd, c0, rd, c0+1, nrows,
                    "Gore Point vs Speed", "Avg Speed (kmph)", YELLOW, w=14, h=11)
    ws_d.add_chart(ch8, "S121")

print("  Charts embedded!")

# ============================================================
# INSIGHTS ROW on Dashboard
# ============================================================
row_ins = 157
for cc in range(1, 36):
    ws_d.cell(row=row_ins, column=cc).fill = fill(HEADER)
merge_cell(ws_d, row_ins, 1, row_ins, 35,
    "  KEY INSIGHTS & POLICY RECOMMENDATIONS",
    bg=HEADER,
    font=ft(bold=True, size=13, color=TEAL),
    align=al('left', 'center'))
ws_d.row_dimensions[row_ins].height = 24

insights = [
    (BLUE,   f"Peak hour is {peak_hr}:00",
              f"Segment '{slowest_seg}' is slowest at {speed_seg.iloc[0]['Avg Speed (kmph)']:.1f} kmph — deploy adaptive signals at peak hour"),
    (RED,    f"High congestion zones: {dense_seg}",
              f"Deploy sensor-based merge management at '{dense_seg}' — highest avg traffic density"),
    (TEAL,   f"50% EV adoption => {co2_drop50:.1f}% CO2 reduction",
              f"City should mandate EV incentives; switching half the fleet could cut emissions by {co2_drop50:.1f}%"),
    (PURPLE, f"Best efficiency: Segment '{top_tei_seg}' (TEI={tei_rank.iloc[0]['Avg TEI']:.3f})",
              f"Study signal timing + lane config of '{top_tei_seg}' as model for underperforming segments"),
]

for idx, (accent, header_text, detail) in enumerate(insights):
    rr = row_ins + 1 + idx * 4
    ws_d.row_dimensions[rr].height   = 7
    ws_d.row_dimensions[rr+1].height = 18
    ws_d.row_dimensions[rr+2].height = 16
    ws_d.row_dimensions[rr+3].height = 5
    for cc in range(1, 36):
        ws_d.cell(row=rr, column=cc).fill = fill(accent)
    merge_cell(ws_d, rr+1, 1, rr+1, 35, f"  {header_text}",
               bg=CARD, font=ft(bold=True, size=11, color="FFFFFF"), align=al('left', 'center'))
    merge_cell(ws_d, rr+2, 1, rr+2, 35, f"      -> {detail}",
               bg=BG, font=ft(size=9, italic=True, color=GREY), align=al('left', 'center'))
    for cc in range(1, 36):
        ws_d.cell(row=rr+3, column=cc).fill = fill(BG)

# ============================================================
# SHEET C: CLEANED DATA
# ============================================================
print("\n[6/6] Writing cleaned data sheet...")
ws_c = wb.create_sheet("Cleaned Data")

cols = list(df.columns)
col_widths_map = {
    'timestamp': 20, 'day of week': 13, 'road segment id': 16,
    'x-coord': 9, 'y-coord': 9, 'vehicles': 10, 'lane count': 11,
    'heavy vehicle pct': 17, 'traffic density': 15, 'rain mm': 10,
    'temperature C': 14, 'visibility m': 14, 'road wetness index': 18,
    'ev percentage': 14, 'fuel mix index': 14, 'avg speed kmph': 16,
    'gore point': 11, 'distance to gore m': 18, 'CO2 emission': 14,
    'NOx': 9, 'PM2.5': 9, 'congestion': 12, 'hour': 7,
    'traffic efficiency index': 24, 'rain_bin': 16,
}

# Header
for ci, col in enumerate(cols):
    hc = ws_c.cell(row=1, column=ci+1, value=col)
    hc.fill = fill(HEADER)
    hc.font = Font(bold=True, size=10, color=TEAL, name="Calibri")
    hc.alignment = Alignment(horizontal='center', vertical='center')
    ws_c.column_dimensions[get_column_letter(ci+1)].width = col_widths_map.get(col, 14)

ws_c.freeze_panes = "A2"
ws_c.row_dimensions[1].height = 18

# Data rows
for ri, row_vals in enumerate(df.values):
    rnum = ri + 2
    bg_row = DARK2 if ri % 2 == 0 else BG
    for ci, val in enumerate(row_vals):
        if pd.isna(val):
            cell_val = ""
        elif isinstance(val, (float, np.floating)):
            cell_val = round(float(val), 4)
        else:
            cell_val = val
        dc = ws_c.cell(row=rnum, column=ci+1, value=cell_val)
        dc.fill      = fill(bg_row)
        dc.font      = Font(size=9, color="CCCCCC", name="Calibri")
        dc.alignment = Alignment(horizontal='center', vertical='center')

# ============================================================
# SHEET D: KPI SUMMARY
# ============================================================
ws_k = wb.create_sheet("KPI Summary")
for r in range(1, 80):
    for c in range(1, 20):
        ws_k.cell(row=r, column=c).fill = fill(BG)

merge_cell(ws_k, 1, 1, 2, 16,
    "KEY PERFORMANCE INDICATORS  |  Highway Traffic Analysis  |  SRIJAN '26",
    bg=HEADER, font=ft(bold=True, size=14, color=TEAL), align=al('center'))
ws_k.row_dimensions[1].height = 24
ws_k.row_dimensions[2].height = 24

kpi_list = [
    ("Total Traffic Volume",       f"{total_vol:,}",            "Sum of all vehicle counts across clean records",    BLUE),
    ("Average Speed (all segs)",   f"{avg_spd:.2f} kmph",       "Mean avg speed across all 4012 records",            TEAL),
    ("Average CO2 Emission",       f"{avg_co2:.2f}",            "Mean CO2 across all records",                       RED),
    ("Emission per Vehicle",       f"{em_per_veh:.5f}",         "CO2 / vehicle ratio",                              YELLOW),
    ("Avg Traffic Eff. Index",     f"{avg_tei:.4f}",            "Speed / Density — higher = more efficient",        PURPLE),
    ("Total NOx",                  f"{total_nox:.1f}",          "Sum of NOx across all records",                    "FF8C42"),
    ("Total PM2.5",                f"{total_pm25:.2f}",         "Sum of PM2.5 across all records",                  "42A5F5"),
    ("Peak Traffic Hour",          f"{peak_hr}:00",             "Hour with highest cumulative vehicle count",         BLUE),
    ("Slowest Road Segment",       slowest_seg,                 f"{speed_seg.iloc[0]['Avg Speed (kmph)']:.1f} kmph average speed", RED),
    ("Most Congested Segment",     dense_seg,                   f"{density_seg.iloc[0]['Avg Density']:.1f} avg traffic density",   RED),
    ("Most Efficient Segment",     top_tei_seg,                 f"TEI = {tei_rank.iloc[0]['Avg TEI']:.4f}",          TEAL),
    ("CO2 Cut at 50% EV",          f"{co2_drop50:.1f}%",        "Projected CO2 reduction if 50% fleet goes EV",     TEAL),
    ("Dataset (clean rows)",       f"{len(df):,}",              f"After removing {raw_count - len(df)} exact duplicates", PURPLE),
]

for ci in range(1, 17):
    ws_k.column_dimensions[get_column_letter(ci)].width = 28

accent_cycle = [BLUE, TEAL, RED, YELLOW, PURPLE, "FF8C42", "42A5F5",
                "EF5350", "AB47BC", "26A69A", "FFCA28", "66BB6A", PURPLE]

for idx, (kpi_name, kpi_val, kpi_desc, accent_c) in enumerate(kpi_list):
    rr = 4 + idx * 5
    ws_k.row_dimensions[rr].height   = 6
    ws_k.row_dimensions[rr+1].height = 18
    ws_k.row_dimensions[rr+2].height = 22
    ws_k.row_dimensions[rr+3].height = 15
    ws_k.row_dimensions[rr+4].height = 6
    for cc in range(1, 17):
        ws_k.cell(row=rr, column=cc).fill = fill(accent_c)
    merge_cell(ws_k, rr+1, 1, rr+1, 8, kpi_name,
               bg=CARD, font=ft(bold=True, size=11, color=GREY), align=al('left', 'center'))
    merge_cell(ws_k, rr+2, 1, rr+2, 4, kpi_val,
               bg=CARD, font=ft(bold=True, size=16, color=accent_c), align=al('left', 'center'))
    merge_cell(ws_k, rr+2, 5, rr+2, 8, kpi_desc,
               bg=CARD, font=ft(size=9, italic=True, color=GREY), align=al('left', 'center'))
    for cc in range(1, 17):
        ws_k.cell(row=rr+3, column=cc).fill = fill(BG)
        ws_k.cell(row=rr+4, column=cc).fill = fill(BG)

# ============================================================
# SET SHEET ORDER & SAVE
# ============================================================
wb.active = ws_d   # Dashboard first

output = r'c:\Users\name\Desktop\PROJECTS\GLOSA-BHARAT-main\Highway_Traffic_Dashboard.xlsx'
wb.save(output)

print(f"\n{'='*60}")
print(f"  DASHBOARD SAVED!")
print(f"{'='*60}")
print(f"  File  : Highway_Traffic_Dashboard.xlsx")
print(f"  Sheets: Dashboard | KPI Summary | Analysis Data | Cleaned Data")
print(f"  Charts: 8 charts embedded")
print(f"  Rows  : {len(df):,} clean  (removed {raw_count-len(df)} duplicates)")
print(f"  Path  : {output}")
print(f"{'='*60}")
